#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
It contains solutions that make Excel's work easier.
"""

__author__ = 'ibrahim CÖRÜT'
__email__ = 'ibrhmcorut@gmail.com'

import os
from copy import copy
from matplotlib.colors import cnames
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.copier import WorksheetCopy
from openpyxl.workbook.workbook import save_workbook
from ..Decorators import try_except
from ..Shell import print_error


def get_color(color):
    try:
        if str(color).startswith('#'):
            return Color(str(color).strip('#'))
        else:
            return cnames[color.lower()].strip('#')
    except KeyError:
        return Color()


class Excel:
    def __init__(self, book=None, sheet=None):
        self.book = book
        self.sheet = sheet
        self.work_book = None
        self.work_sheet = None
        self.open(book, sheet)

    def open(self, book=None, sheet=None):
        """
        Used to create excel or read existing excel
        :param book: Destination excel file path.
        :param sheet: Target sheet name
        """
        if book:
            self.book = book
        if sheet:
            self.sheet = sheet
        try:
            if book:
                if not os.path.exists(book):
                    wb = Workbook(book)
                    wb.create_sheet(sheet)
                    wb.save(book)
                    wb.close()
                self.work_book = load_workbook(book)
            if sheet:
                try:
                    self.work_sheet = self.work_book[sheet]
                except KeyError:
                    self.work_book.create_sheet(sheet)
                    self.work_sheet = self.work_book[sheet]
            if self.work_book:
                self.work_book.save(self.book)
        except Exception as error:
            print_error(error, locals())
        print(f'Excel Open Success ---> Book:{self.book} ---> Sheet:{self.sheet}')
        return self

    @try_except
    def close(self):
        """
        Closes the Excel file
        """
        save_workbook(self.work_book, self.book)
        self.work_book.close()
        self.work_book = None
        self.work_sheet = None
        print(f'Excel Close Success ---> Book:{self.book}')

    def write(
            self, row=1, column=1, data=None, color=None,
            border=False, bold=False, size=8, alignments=None,
            multiple_columns_in_row=False, multiple_columns_and_rows=False,
            color_by_word_inside=False
    ):
        """

        :param row:
        :param column:
        :param data:
        :param color:
        :param border:
        :param bold:
        :param size:
        :param alignments:
            horizontal_alignments:
                "general", "left", "center", "right", "fill", "justify",
                "centerContinuous", "distributed"
            vertical_alignments:
                top", "center", "bottom", "justify", "distributed"
            for example:
                {'horizontal': 'right', 'vertical': 'center'}
        :param multiple_columns_in_row:
        :param multiple_columns_and_rows:
        :param color_by_word_inside:
        """
        try:
            if alignments is None:
                alignments = {'horizontal': 'right', 'vertical': 'center'}
            else:
                alignments.update(alignments)
            color_temp = color
            if multiple_columns_in_row:
                for i, j in enumerate(data):
                    if color_by_word_inside and color is None and str(j).lower() in cnames.keys():
                        color_temp = j
                    if color_temp:
                        new = self.work_sheet.cell(row=row, column=column + i, value=j)
                        new.fill = PatternFill(fgColor=get_color(color_temp), fill_type='solid')
                    else:
                        new = self.work_sheet.cell(row=row, column=column + i, value=j)
                    if border:
                        thin = Side(border_style="thin", color="000000")
                        new.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    new.font = Font(bold=bold, size=size)
                    new.alignment = Alignment(**alignments)
            elif multiple_columns_and_rows:
                for i, rows_data in enumerate(data):
                    for j, rows_value in enumerate(rows_data):
                        if (
                                color_by_word_inside and
                                color is None and
                                str(rows_value).lower() in cnames.keys()
                        ):
                            color_temp = rows_value
                        if color_temp:
                            new = self.work_sheet.cell(
                                row=row+i, column=column+j, value=rows_value
                            )
                            new.fill = PatternFill(
                                fgColor=get_color(color_temp), fill_type='solid'
                            )
                        else:
                            new = self.work_sheet.cell(
                                row=row+i, column=column+j, value=rows_value
                            )
                        if border:
                            thin = Side(border_style="thin", color="000000")
                            new.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        new.font = Font(bold=bold, size=size)
                        new.alignment = Alignment(**alignments)
            else:
                new = self.work_sheet.cell(row=row, column=column, value=data)
                if color_by_word_inside and color is None and str(data).lower() in cnames.keys():
                    color_temp = data
                if color_temp:
                    new.fill = PatternFill(fgColor=get_color(color_temp), fill_type='solid')
                if border:
                    thin = Side(border_style="thin", color="000000")
                    new.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                new.font = Font(bold=bold, size=size)
                new.alignment = Alignment(**alignments)
            for column in self.work_sheet.columns:
                length = max(len(str(cell.value)) for cell in column)
                for check in range(len(column)):
                    try:
                        self.work_sheet.column_dimensions[column[check]].width = length + 2
                        break
                    except TypeError as error:
                        pass
            self.work_book.save(self.book)
            print(f'Excel All Data Write Success ---> Sheet:{self.sheet} ---> Data:{data}')
        except Exception as error:
            print_error(error, locals())

    def merge_cells(
            self, start_row, start_column, end_row, end_column, value,
            color=None, border=True, bold=True, size=8
    ):
        try:
            self.work_sheet.merge_cells(
                start_row=start_row, start_column=start_column,
                end_row=end_row, end_column=end_column
            )
            new = self.work_sheet.cell(row=start_row, column=start_column, value=value)
            if color:
                new.fill = PatternFill(fgColor=get_color(color), fill_type='solid')
            if border:
                thin = Side(border_style="thin", color="000000")
                new.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            new.font = Font(bold=bold, size=size)
            new.alignment = Alignment(horizontal='center', vertical='center')
            self.work_sheet.column_dimensions[get_column_letter(start_column)].auto_size = True
            self.work_book.save(self.book)
            print(f'Excel Set Merge Cell ---> Sheet:{self.sheet}')
            print(f'Start Row:{start_row} - Start Column:{start_column}')
            print(f'End   Row:{end_row}   - End   Column:{end_column}')
        except Exception as error:
            print_error(error, locals())

    def merge_cells_remove(self, start_row, start_column, end_row, end_column):
        try:
            self.work_sheet.unmerge_cells(
                start_row=start_row, start_column=start_column,
                end_row=end_row, end_column=end_column
            )
            self.work_book.save(self.book)
            print(f'Excel Set Merge Cell Remove ---> Sheet:{self.sheet}')
            print(f'Start Row:{start_row} - Start Column:{start_column}')
            print(f'End   Row:{end_row}   - End   Column:{end_column}')
        except Exception as error:
            print_error(error, locals())

    def read(self, row, column):
        data = None
        try:
            data = self.work_sheet.cell(row=row, column=column).value
            print(f'Excel Read > Sheet:{self.sheet} Line:{row} Column:{column} ---> Data:{data}')
        except Exception as error:
            print_error(error, locals())
        return data

    def read_all_sheet(self, row_start=1, column_start=1, row_end=None, column_end=None):
        data = []
        try:
            if row_end is None:
                row_end = self.work_sheet.max_row
            if column_end is None:
                column_end = self.work_sheet.max_column
            for i in range(row_start, row_end + 1):
                line = []
                for j in range(column_start, column_end):
                    line.append(self.work_sheet.cell(row=i, column=j).value)
                data.append(line)
            print(f'Excel Read All Sheet Success ---> Sheet:{self.sheet} Data:{data}')
        except Exception as error:
            print_error(error, locals())
        return data

    def get_last_line(self, line_status='row'):
        """
        Returns the last row or last column number in the target excel and target sheet
        :param line_status: 'row' or 'column' option
        :return: Returns integer value
        """
        data = 0
        try:
            line_status = line_status.title()
            if line_status == 'Row':
                data = self.work_sheet.max_row
            elif line_status == 'Column':
                data = self.work_sheet.max_column
            print(f'Excel Get Last Row/Column from Sheet:{self.sheet} ---> {line_status}:{data}')
        except Exception as error:
            print_error(error, locals())
        return data

    @try_except
    def get_sheet_names(self):
        sheets = self.work_book.sheetnames
        print(f'Excel Get Sheet Names:{sheets}')
        return sheets

    def column_clone(self, column, column_to_be_copied):
        try:
            self.work_sheet.insert_cols(column)
            for row in range(1, self.work_sheet.max_row + 1):
                x = self.work_sheet.cell(row=row, column=column_to_be_copied)
                y = self.work_sheet.cell(row=row, column=column)
                # y.value = copy(x.value)
                y.style = copy(x.style)
                y.font = copy(x.font)
                y.border = copy(x.border)
                y.fill = copy(x.fill)
                y.number_format = copy(x.number_format)
                y.protection = copy(x.protection)
                y.alignment = copy(x.alignment)
            self.work_book.save(self.book)
            print(f'Excel Column Clone Success ---> Sheet:{self.sheet} Column:{column}')
        except Exception as error:
            print_error(error, locals())

    @staticmethod
    def sheet_copy_only_data(book, sheet, delete_old_sheet=False):
        try:
            workbook = load_workbook(book)
            worksheet = workbook[sheet]
            new_worksheet = workbook.create_sheet('New_Sheet_Name')
            instance = WorksheetCopy(worksheet, new_worksheet)
            WorksheetCopy.copy_worksheet(instance)
            if delete_old_sheet:
                workbook.remove(worksheet)
                worksheet_update = workbook['New_Sheet_Name']
                worksheet_update.title = sheet
            workbook.save(book)
            workbook.close()
            print(f'Excel Sheet Copy Success ---> Sheet:{sheet}')
        except Exception as error:
            print_error(error, locals())

    def column_delete(self, column):
        try:
            self.work_sheet.delete_cols(column)
            self.work_book.save(self.book)
            print(f'Excel Column Delete Success ---> Sheet:{self.sheet} Column:{column}')
        except Exception as error:
            print_error(error, locals())

    def add_chart_on_sheet(
            self,
            title_min_col, title_min_row, title_max_col, title_max_row,
            value_min_col, value_min_row, value_max_row,
            width=1,
            title=None, x_title=None, y_title=None,
            chart_position_show='A1'
    ):
        try:
            worksheet = self.work_book[self.work_sheet.title]
            values = Reference(
                worksheet, min_col=title_min_col, min_row=title_min_row,
                max_col=title_max_col, max_row=title_max_row
            )
            cats = Reference(
                worksheet, min_col=value_min_col, min_row=value_min_row, max_row=value_max_row
            )
            chart = BarChart()
            chart.add_data(values, titles_from_data=True)
            chart.set_categories(cats)
            if title is not None:
                chart.title = title
            if x_title is not None:
                chart.x_axis.title = x_title
            if y_title is not None:
                chart.y_axis.title = y_title
            chart.width = width
            worksheet.add_chart(chart, chart_position_show)
            self.work_book.save(self.book)
            print(f'Excel Add Chart Success ---> Book:{self.book} ---> Sheet:{self.sheet.title}')
        except Exception as error:
            print_error(error, locals())
