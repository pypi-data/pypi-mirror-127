from pathlib import Path
from typing import Type, Union

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl_image_loader import SheetImageLoader

from .schemas import ImageCell, SheetTemplate


def validate_xlsx(
        fp: Union[str, Path], model: Type[SheetTemplate],
        sheet_index: int = 0, ignore_validate_errors=False
):
    wb = load_workbook(fp)
    sheet = wb.worksheets[sheet_index]
    rows = sheet.iter_rows()
    headers = {th.value: i for i, th in enumerate(next(rows))}
    image_loader = SheetImageLoader(sheet)

    img_cols = tuple(
        (n, headers[f.alias]) for n, f in model.__fields__.items()
        if issubclass(f.type_, ImageCell) and f.alias in headers
    )
    for row_x, raw_row in enumerate(rows, start=2):
        try:
            item = model.validate({key: v for key, v in zip(headers, raw_row) if key is not None})
        except BaseException as e:
            if ignore_validate_errors:
                continue
            raise e
        for field, col in img_cols:
            if image_loader.image_in(coord := f'{get_column_letter(col + 1)}{row_x}'):
                setattr(item, field, image_loader.get(coord))
        yield item


if __name__ == 'main':
    pass
