from __future__ import annotations

from pathlib import Path
import shutil
from typing import Literal

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
import pandas as pd

from assistants.utility import excel_interface

ADULT_INFO_FIELDS = ["membership_number", "known_as", "surname", "email", "phone_number", "country", "region", "county", "district", "scout_group"]


def compliance_supporting_data(data: pd.DataFrame, region: str) -> None:
    date = data.attrs["report_date"]
    print(f"CSD - {date}")

    # data sources (map of excel sheet/tab name -> dataframe)
    sources: dict[str, pd.DataFrame] = {
        "Appointments": data,
        "Persistent": adults_by_type_overdue(data, ["Safety", "Safeguarding"]),  # persistently
        "No Email": _no_email(data),  # No email address
    }

    template_report = Path("data/compliance-supporting-data/csd-county-template.xlsx")
    counties = sorted(set(data["county"].dropna().array.to_numpy())) + ["Region"]
    for county in counties:
        is_region_team = county == "Region"
        print(f"CSD - {county} starting")
        report_path = template_report.parent / region / county / f"csd - {date} - {county}.xlsx"
        report_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy(template_report, report_path)
        with excel_interface.load_workbook(report_path) as xl:
            for sheet_name, source in sources.items():
                mask = source["county"].isna() if is_region_team else source["county"] == county
                source.loc[mask].to_excel(xl, sheet_name, startrow=2, header=False, index=False)
                # hidden_columns={"Appointments": r[35:161 + 1], },


RULE_OVERDUE = CellIsRule(operator="between", formula=["1", "EDATE(TODAY(),-3)"], fill=PatternFill(end_color="FB9797"))
RULE_SOON = CellIsRule(operator="between", formula=["TODAY()", "EDATE(TODAY(),1)"], fill=PatternFill(end_color="FCD380"))


def by_adult_overview(data: pd.DataFrame, file_prefix: str, types: tuple[str, ...] = ("Safety", "Safeguarding")) -> None:
    """Create an overview of the adults at various statuses.

    Saves one CSV file per compliance status, and generates an Excel workbook
    with one worksheet per compliance status, including conditional formatting
    and frozen panes.

    Args:
        data: Compliance report to generate lists for
        file_prefix:
            File prefix to use when saving files. Can include a path prefix,
            for example "path/to/reports/folder/file_prefix "
        types: compliance types to run overview for. Default is Safety and
            Safeguarding

    Returns:
        None

    """
    # ensure output directory exists
    Path(f"{file_prefix}_.tmp").parent.mkdir(parents=True, exist_ok=True)

    date = data.attrs["report_date"]
    out = adults_by_type_overdue(data, [*types], {"FirstAid": "First Aid"})
    out = out.rename(columns={"Safety Due By": "Safety OD Since", "Safeguarding Due By": "Safeguarding OD Since", "Firstaid Due By": "First Aid OD Since"})
    out.to_csv(f"{file_prefix} Overdue Adults - {date}.csv", encoding="utf-8-sig", index=False)

    soon = adults_by_type_due_soon(data, [*types], {"FirstAid": "First Aid"})
    soon.to_csv(f"{file_prefix} Due Soon Adults - {date}.csv", encoding="utf-8-sig", index=False)

    with pd.ExcelWriter(f"{file_prefix} MOGL by adult - {date}.xlsx", engine="openpyxl") as writer:
        writer.datetime_format = "DD/MM/YYYY"  # bug in OpenpxylWriter.__init__ not passing this to super()
        out.to_excel(writer, sheet_name=f"Overdue Adults - {date}", index=False, freeze_panes=(1, 0))
        soon.to_excel(writer, sheet_name=f"Due Soon Adults - {date}", index=False, freeze_panes=(1, 0))

        # Conditional formatting
        wb = writer.book  # NoQA
        wb[f"Overdue Adults - {date}"].conditional_formatting.add("J1:K9999", RULE_OVERDUE)
        wb[f"Due Soon Adults - {date}"].conditional_formatting.add("J1:K9999", RULE_SOON)


def adults_by_type_overdue(
    extract_data: pd.DataFrame,
    types: list[str],
    types_map: dict[str, str] | None = None,
) -> pd.DataFrame:
    return _adults_by_type_status(extract_data, types, types_map or {}, "overdue")


def adults_by_type_due_soon(
    extract_data: pd.DataFrame,
    types: list[str],
    types_map: dict[str, str] | None = None,
) -> pd.DataFrame:
    return _adults_by_type_status(extract_data, types, types_map or {}, "soon")


def _adults_by_type_status(
    extract_data: pd.DataFrame,
    types: list[str],
    types_map: dict[str, str],
    status: Literal["not_req", "overdue", "soon", "start", "valid"],
) -> pd.DataFrame:
    extract_data = extract_data.copy()  # fix set with copy warning (we assign to the original data when adding overdue since dates)
    names = [types_map.get(t, t) for t in types]
    cols = [f"ByAdult{t}" for t in types]

    # Get boolean matrix of overdue or not by adult, and a scalar per adult if any are overdue
    mask_arr = extract_data[cols] == status
    mask_adults = mask_arr.any(axis=1)

    # add overdue since dates, if they exist
    g = extract_data.groupby("membership_number")
    _due_by_cols = []
    for t in types:
        if f"DueBy{t}" not in extract_data.columns:
            continue
        _due_by_cols.append(f"{t.lower()}_due_by")
        # Min non-compliant since//due by date by member
        extract_data[f"{t.lower()}_due_by"] = g[f"DueBy{t}"].transform("min")
        # Blank out compliant values
        extract_data.loc[~mask_arr[f"ByAdult{t}"], f"{t.lower()}_due_by"] = float("NaN")

    # Create output array
    out = extract_data.loc[mask_adults, ADULT_INFO_FIELDS + _due_by_cols].sort_values("membership_number").reset_index(drop=True)

    # Map true values to the names list
    mapped = mask_arr[mask_adults] * names

    # Join the columns, ignoring blank strings
    out["type"] = [", ".join(filter(None, row)) for row in zip(*(mapped.iloc[:, i] for i in range(len(cols))))]

    # Sort for presentation
    out = out.sort_values(["district", "scout_group"], kind="stable")
    out["district"] = out["district"].fillna("County")

    # Rename for export
    out = out.rename(columns=lambda field: field.replace("_", " ").title())
    return out


def _no_email(data: pd.DataFrame) -> pd.DataFrame:
    """No email address."""
    return data.loc[data["email"].isna(), ADULT_INFO_FIELDS].drop_duplicates(subset={"membership_number"})
